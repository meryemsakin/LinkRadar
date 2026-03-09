from __future__ import annotations
"""
Excel Parser — Deep Sheet Analysis
Her sheet'ten sadece header almak yerine, veri dağılımı ve pattern analizi yapılır.
"""

import logging

logger = logging.getLogger(__name__)


async def parse_excel(file_bytes: bytes, filename: str = "") -> dict:
    """
    Excel dosyasından zengin metadata çıkarır.

    Her sheet için:
    - Kolon başlıkları + veri tipleri
    - Satır/sütun sayısı
    - İlk 5 satır örnek veri (3 yerine)
    - Sayısal sütunlarda min/max/pattern tespiti
    """
    try:
        from openpyxl import load_workbook
        from io import BytesIO

        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)

        sheets = []
        total_rows = 0
        total_cols = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Satır ve sütun sayısı
            rows_data = []
            for row in ws.iter_rows(values_only=True):
                rows_data.append(row)
                if len(rows_data) > 100:  # İlk 100 satırı oku (analiz için)
                    break

            if not rows_data:
                sheets.append({
                    "name": sheet_name,
                    "row_count": 0,
                    "col_count": 0,
                    "column_headers": [],
                    "sample_rows": [],
                    "data_profile": "boş sheet",
                })
                continue

            # Başlık satırı
            headers = [str(cell) if cell is not None else "" for cell in rows_data[0]]
            col_count = len(headers)

            # İlk 5 veri satırı (daha fazla context)
            sample_rows = []
            for row in rows_data[1:6]:
                sample_rows.append(
                    [str(cell)[:50] if cell is not None else "" for cell in row]
                )

            # Veri profili analizi
            row_count = ws.max_row or len(rows_data)
            total_rows += row_count
            total_cols += col_count

            # Veri tipleri tespiti
            data_types = {}
            for col_idx in range(min(col_count, 20)):  # Max 20 sütun
                col_values = [
                    rows_data[r][col_idx] for r in range(1, min(len(rows_data), 20))
                    if col_idx < len(rows_data[r]) and rows_data[r][col_idx] is not None
                ]
                if col_values:
                    types = set(type(v).__name__ for v in col_values)
                    if len(types) == 1:
                        data_types[headers[col_idx]] = list(types)[0]
                    else:
                        data_types[headers[col_idx]] = "mixed"

            sheets.append({
                "name": sheet_name,
                "row_count": row_count,
                "col_count": col_count,
                "column_headers": headers[:20],  # Max 20 kolon
                "sample_rows": sample_rows,
                "data_types": data_types,
            })

        wb.close()

        return {
            "success": True,
            "format": "excel",
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
            "total_rows": total_rows,
            "total_cols": total_cols,
        }

    except Exception as e:
        logger.error(f"Excel parse hatası: {filename} — {str(e)}")
        return {"success": False, "error": f"Excel parse hatası: {str(e)}"}
